from openpyxl import load_workbook

def read_signup_test_data(file_path: str, sheet_name: str = "sign-up"):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data